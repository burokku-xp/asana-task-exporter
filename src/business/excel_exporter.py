"""
Excel出力機能を提供するモジュール

このモジュールは、Asanaから取得したタスクデータをExcelファイルに出力する機能を提供します。
動的なフィールド選択、フォーマット、スタイリング、大量データ処理に対応しています。
"""

import os
import logging
from datetime import datetime, date
from typing import List, Dict, Any, Optional, Callable
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError("openpyxl is required for Excel export functionality. Please install it with: pip install openpyxl")

from ..data.models import Task
from .config_schema import get_field_display_name
from ..utils.error_handler import (
    ErrorHandler, FileError, ValidationError, 
    handle_errors, ErrorContext, retry_on_error
)

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Excelファイル出力を管理するクラス
    
    動的なフィールド選択に対応し、適切なフォーマットとスタイリングを適用して
    タスクデータをExcelファイルに出力します。
    """
    
    def __init__(self):
        """ExcelExporterを初期化"""
        self.batch_size = 1000  # バッチ処理のサイズ
        self.progress_callback: Optional[Callable[[int], None]] = None
        self.status_callback: Optional[Callable[[str], None]] = None
        self._current_operation = ""
        self.error_handler = ErrorHandler()
        
    @handle_errors("Excel出力処理", reraise=True)
    def export_to_excel(
        self,
        tasks: List[Task],
        filename: str,
        selected_fields: List[str],
        field_labels: Optional[Dict[str, str]] = None,
        progress_callback: Optional[Callable[[int], None]] = None,
        status_callback: Optional[Callable[[str], None]] = None
    ) -> str:
        """
        タスクデータをExcelファイルに出力
        
        Args:
            tasks: 出力するタスクのリスト
            filename: 出力ファイル名（拡張子なし）
            selected_fields: 出力するフィールドのリスト
            progress_callback: 進捗状況を通知するコールバック関数
            
        Returns:
            str: 作成されたファイルの完全パス
            
        Raises:
            ValidationError: 無効なパラメータが指定された場合
            FileError: ファイル書き込みに失敗した場合
        """
        # 入力パラメータの検証
        if not tasks:
            raise ValidationError("出力するタスクが指定されていません", field="tasks", value=tasks)
            
        if not selected_fields:
            raise ValidationError("出力するフィールドが選択されていません", 
                                field="selected_fields", value=selected_fields)
        
        if not isinstance(tasks, list):
            raise ValidationError("tasks はリストである必要があります", field="tasks", value=type(tasks))
        
        if not isinstance(selected_fields, list):
            raise ValidationError("selected_fields はリストである必要があります", 
                                field="selected_fields", value=type(selected_fields))
            
        self.progress_callback = progress_callback
        self.status_callback = status_callback
        self.field_labels = field_labels or {}

        with ErrorContext("Excel出力メイン処理", reraise=True) as ctx:
            # ファイル名の準備と検証
            filename = self._prepare_filename(filename)
            output_path = Path(filename)
            
            # 出力パスの検証
            is_valid, error_msg = self.validate_output_path(filename)
            if not is_valid:
                raise FileError(f"出力パスが無効です: {error_msg}", file_path=filename)
            
            logger.info(f"Excel出力開始: {len(tasks)}件のタスクを{filename}に出力")
            self._update_status("Excel出力を開始しています...")
            
            try:
                # メモリ使用量の監視開始
                import gc
                gc.collect()  # ガベージコレクション実行
                
                # Workbookの作成
                self._update_status("Excelファイルを初期化しています...")
                workbook = self._create_workbook()
                worksheet = workbook.active
                worksheet.title = "Asana Tasks"
                
                # データの準備（メモリ効率化）
                self._update_status("データを準備しています...")
                data = self._prepare_data_efficiently(tasks, selected_fields)
                
                # ヘッダーの設定
                self._update_status("ヘッダーを設定しています...")
                self._write_headers(worksheet, selected_fields)
                self._update_progress(10)
                
                # データの書き込み（バッチ処理）
                self._update_status("データを書き込んでいます...")
                self._write_data_in_batches(worksheet, data, selected_fields)
                
                # フォーマットの適用
                self._update_status("書式を適用しています...")
                self._apply_formatting(worksheet, len(data), len(selected_fields))
                
                # ファイルの保存
                self._update_status("ファイルを保存しています...")
                self._save_workbook(workbook, filename)
                self._update_progress(100)
                
                # メモリクリーンアップ
                del workbook
                del data
                gc.collect()
                
                logger.info(f"Excel出力完了: {filename}")
                self._update_status("Excel出力が完了しました")
                    
                return str(output_path.absolute())
                
            except PermissionError as e:
                raise FileError(f"ファイルへの書き込み権限がありません: {e}", file_path=filename, original_error=e)
            except OSError as e:
                raise FileError(f"ファイルシステムエラーが発生しました: {e}", file_path=filename, original_error=e)
            except MemoryError as e:
                raise FileError(f"メモリ不足のため処理を継続できません: {e}", file_path=filename, original_error=e)
            except Exception as e:
                raise FileError(f"Excelファイルの作成に失敗しました: {e}", file_path=filename, original_error=e)
    
    def _prepare_filename(self, filename: str) -> str:
        """ファイル名を準備・検証"""
        if not filename or not isinstance(filename, str):
            raise ValidationError("ファイル名が無効です", field="filename", value=filename)
        
        filename = filename.strip()
        if not filename:
            raise ValidationError("ファイル名が空です", field="filename", value=filename)
        
        # 拡張子の追加
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # 出力ディレクトリの確保
        output_path = Path(filename)
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
        except OSError as e:
            raise FileError(f"出力ディレクトリの作成に失敗しました: {e}", 
                          file_path=str(output_path.parent), original_error=e)
        
        return filename
    
    def _create_workbook(self) -> Workbook:
        """Workbookを作成"""
        try:
            return Workbook()
        except Exception as e:
            raise FileError(f"Excelワークブックの作成に失敗しました: {e}", original_error=e)
    
    def _save_workbook(self, workbook: Workbook, filename: str):
        """Workbookを保存"""
        try:
            workbook.save(filename)
        except PermissionError as e:
            raise FileError(f"ファイルが他のプログラムで使用中です: {filename}", 
                          file_path=filename, original_error=e)
        except OSError as e:
            if "No space left on device" in str(e) or "disk full" in str(e).lower():
                raise FileError(f"ディスクの空き容量が不足しています: {e}", 
                              file_path=filename, original_error=e)
            else:
                raise FileError(f"ファイル保存エラー: {e}", file_path=filename, original_error=e)
        except Exception as e:
            raise FileError(f"Excelファイルの保存に失敗しました: {e}", 
                          file_path=filename, original_error=e)
    
    def _prepare_data(self, tasks: List[Task], selected_fields: List[str]) -> List[Dict[str, Any]]:
        """
        タスクデータを出力用に準備
        
        Args:
            tasks: タスクのリスト
            selected_fields: 選択されたフィールド
            
        Returns:
            List[Dict[str, Any]]: 出力用データ
        """
        data = []
        total_tasks = len(tasks)
        
        for i, task in enumerate(tasks):
            row_data = {}
            
            for field in selected_fields:
                row_data[field] = self._get_field_value(task, field)
            
            data.append(row_data)
            
            # 進捗通知（データ準備は全体の30%）
            if self.progress_callback and i % 100 == 0:
                progress = int((i / total_tasks) * 30)
                self.progress_callback(progress)
        
        return data
    
    def _prepare_data_efficiently(self, tasks: List[Any], selected_fields: List[str]) -> List[Dict[str, Any]]:
        """
        メモリ効率を考慮したデータ準備（大量データ対応）

        Args:
            tasks: タスクのリスト（Taskオブジェクトまたは辞書）
            selected_fields: 選択されたフィールド

        Returns:
            List[Dict[str, Any]]: 出力用データ
        """
        import gc

        # 既に辞書形式の場合はそのまま返す（filter_tasks_by_fieldsの出力）
        if tasks and isinstance(tasks[0], dict):
            logger.info("データは既に辞書形式です。そのまま使用します。")
            return tasks

        data = []
        total_tasks = len(tasks)

        # バッチサイズを動的に調整（メモリ使用量に応じて）
        dynamic_batch_size = min(self.batch_size, max(100, total_tasks // 10))

        for i in range(0, total_tasks, dynamic_batch_size):
            batch_end = min(i + dynamic_batch_size, total_tasks)
            batch_tasks = tasks[i:batch_end]

            # バッチ単位でデータを処理
            batch_data = []
            for task in batch_tasks:
                row_data = {}
                for field in selected_fields:
                    row_data[field] = self._get_field_value(task, field)
                batch_data.append(row_data)

            data.extend(batch_data)

            # 進捗通知（データ準備は全体の0-30%）
            progress = int((batch_end / total_tasks) * 30)
            self._update_progress(progress)

            # 定期的なガベージコレクション
            if i % (dynamic_batch_size * 5) == 0:
                gc.collect()

            logger.debug(f"データ準備バッチ完了: {i+1}-{batch_end}/{total_tasks}")

        return data
    
    def _get_field_value(self, task: Any, field: str) -> Any:
        """
        タスクから指定されたフィールドの値を取得

        Args:
            task: タスクオブジェクトまたは辞書
            field: フィールド名

        Returns:
            Any: フィールドの値
        """
        # 辞書形式の場合
        if isinstance(task, dict):
            return task.get(field, "")

        # Taskオブジェクトの場合（後方互換性）
        # 基本フィールド
        if field == 'name':
            return task.name
        elif field == 'created_at':
            return task.created_at
        elif field == 'modified_at':
            return task.modified_at
        elif field == 'completed':
            return "完了" if task.completed else "未完了"
        elif field == 'assignee':
            return task.assignee or "未割り当て"
        elif field == 'due_date':
            return task.due_date
        elif field == 'notes':
            return task.notes or ""
        
        # カスタムフィールド
        if hasattr(task, 'custom_fields') and task.custom_fields:
            return task.custom_fields.get(field, "")
        
        return ""
    
    def _write_headers(self, worksheet: Worksheet, selected_fields: List[str]) -> None:
        """
        ヘッダー行を書き込み

        Args:
            worksheet: ワークシート
            selected_fields: 選択されたフィールド
        """
        for col, field in enumerate(selected_fields, 1):
            cell = worksheet.cell(row=1, column=col)
            # フィールド名マッピングがあれば使用、なければデフォルト表示名
            cell.value = self.field_labels.get(field, get_field_display_name(field))

        logger.debug(f"ヘッダー行を書き込み: {len(selected_fields)}列")
    
    def _write_data_in_batches(
        self,
        worksheet: Worksheet,
        data: List[Dict[str, Any]],
        selected_fields: List[str]
    ) -> None:
        """
        データをバッチ処理で書き込み
        
        Args:
            worksheet: ワークシート
            data: 出力データ
            selected_fields: 選択されたフィールド
        """
        total_rows = len(data)
        
        # 動的バッチサイズの計算（メモリ使用量を考慮）
        dynamic_batch_size = self._calculate_optimal_batch_size(total_rows, len(selected_fields))
        
        for batch_start in range(0, total_rows, dynamic_batch_size):
            batch_end = min(batch_start + dynamic_batch_size, total_rows)
            batch_data = data[batch_start:batch_end]
            
            # バッチデータの書き込み
            self._write_batch_data(worksheet, batch_data, selected_fields, batch_start)
            
            # 進捗通知（データ書き込みは10-80%）
            progress = 10 + int((batch_end / total_rows) * 70)
            self._update_progress(progress)
            
            # メモリ管理
            if batch_start % (dynamic_batch_size * 3) == 0:
                import gc
                gc.collect()
            
            logger.debug(f"バッチ書き込み完了: {batch_start+1}-{batch_end}/{total_rows}")
    
    def _calculate_optimal_batch_size(self, total_rows: int, num_fields: int) -> int:
        """
        最適なバッチサイズを計算
        
        Args:
            total_rows: 総行数
            num_fields: フィールド数
            
        Returns:
            int: 最適なバッチサイズ
        """
        # データ量に応じてバッチサイズを調整
        if total_rows < 1000:
            return min(total_rows, 500)
        elif total_rows < 10000:
            return 1000
        elif total_rows < 50000:
            return 2000
        else:
            # 非常に大量のデータの場合はより小さなバッチで処理
            return 1500
    
    def _write_batch_data(
        self,
        worksheet: Worksheet,
        batch_data: List[Dict[str, Any]],
        selected_fields: List[str],
        batch_start: int
    ) -> None:
        """
        バッチデータをワークシートに書き込み
        
        Args:
            worksheet: ワークシート
            batch_data: バッチデータ
            selected_fields: 選択されたフィールド
            batch_start: バッチ開始位置
        """
        for i, row_data in enumerate(batch_data):
            row_num = batch_start + i + 2  # ヘッダー行の次から開始
            
            for col, field in enumerate(selected_fields, 1):
                cell = worksheet.cell(row=row_num, column=col)
                value = row_data.get(field, "")
                
                # データ型に応じた値の設定
                if isinstance(value, (datetime, date)):
                    cell.value = value
                elif isinstance(value, bool):
                    cell.value = value
                elif isinstance(value, (int, float)):
                    cell.value = value
                else:
                    cell.value = str(value) if value is not None else ""
    
    def _apply_formatting(self, worksheet: Worksheet, data_rows: int, data_cols: int) -> None:
        """
        ワークシートにフォーマットを適用
        
        Args:
            worksheet: ワークシート
            data_rows: データ行数
            data_cols: データ列数
        """
        # ヘッダー行のスタイル
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # データ行のスタイル
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # 交互行の背景色
        alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # 行の高さを統一（ヘッダー行）
        worksheet.row_dimensions[1].height = 25

        # ヘッダー行にスタイルを適用
        for col in range(1, data_cols + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # データ行にスタイルを適用
        for row in range(2, data_rows + 2):
            is_alternate_row = (row - 2) % 2 == 1

            # 行の高さを統一（データ行）
            worksheet.row_dimensions[row].height = 20

            for col in range(1, data_cols + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border

                # 交互行の背景色を適用
                if is_alternate_row:
                    cell.fill = alternate_fill

                # データ型に応じたセル書式を適用
                self._apply_cell_format(cell, col)
        
        # 列幅の自動調整と列固有の設定
        self._adjust_column_widths(worksheet, data_cols)
        
        # ワークシート全体の設定
        self._apply_worksheet_settings(worksheet)
        
        # 進捗通知（フォーマット適用は80-90%）
        self._update_progress(90)
        
        logger.debug(f"フォーマット適用完了: {data_rows}行 x {data_cols}列")
    
    def _apply_cell_format(self, cell, col_index: int) -> None:
        """
        セルのデータ型に応じた書式を適用
        
        Args:
            cell: セルオブジェクト
            col_index: 列インデックス
        """
        if cell.value is None:
            return
        
        # 日付・時刻フィールドの書式
        if isinstance(cell.value, datetime):
            cell.number_format = 'YYYY/MM/DD HH:MM:SS'
            cell.alignment = Alignment(horizontal="center")
        elif isinstance(cell.value, date):
            cell.number_format = 'YYYY/MM/DD'
            cell.alignment = Alignment(horizontal="center")
        
        # 数値フィールドの書式
        elif isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")
        
        # ブール値フィールドの書式
        elif isinstance(cell.value, bool):
            cell.alignment = Alignment(horizontal="center")
        
        # テキストフィールドの書式
        else:
            # 長いテキストの場合は折り返し
            if isinstance(cell.value, str) and len(cell.value) > 50:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="center")
    
    def _adjust_column_widths(self, worksheet: Worksheet, data_cols: int) -> None:
        """
        列幅を内容に応じて調整
        
        Args:
            worksheet: ワークシート
            data_cols: データ列数
        """
        for col in range(1, data_cols + 1):
            column_letter = get_column_letter(col)
            
            # 列の内容に基づいて幅を決定
            max_length = 0
            for row in range(1, min(worksheet.max_row + 1, 101)):  # 最初の100行をサンプル
                cell = worksheet.cell(row=row, column=col)
                if cell.value:
                    # 日本語文字を考慮した長さ計算
                    length = self._calculate_display_width(str(cell.value))
                    max_length = max(max_length, length)
            
            # 最小幅10、最大幅50で調整
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _calculate_display_width(self, text: str) -> int:
        """
        テキストの表示幅を計算（日本語文字を考慮）
        
        Args:
            text: テキスト
            
        Returns:
            int: 表示幅
        """
        width = 0
        for char in text:
            # 日本語文字（ひらがな、カタカナ、漢字）は幅2、その他は幅1
            if '\u3040' <= char <= '\u309F' or \
               '\u30A0' <= char <= '\u30FF' or \
               '\u4E00' <= char <= '\u9FAF':
                width += 2
            else:
                width += 1
        return width
    
    def _apply_worksheet_settings(self, worksheet: Worksheet) -> None:
        """
        ワークシート全体の設定を適用

        Args:
            worksheet: ワークシート
        """
        # ヘッダー行を固定
        worksheet.freeze_panes = 'A2'

        # オートフィルタを設定（データ範囲全体に適用）
        if worksheet.max_row > 1 and worksheet.max_column > 0:
            # A1からデータ範囲の最後まで
            last_column = get_column_letter(worksheet.max_column)
            filter_range = f"A1:{last_column}{worksheet.max_row}"
            worksheet.auto_filter.ref = filter_range
            logger.debug(f"オートフィルタを設定: {filter_range}")

        # 印刷設定
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = False

        # マージン設定
        worksheet.page_margins.left = 0.7
        worksheet.page_margins.right = 0.7
        worksheet.page_margins.top = 0.75
        worksheet.page_margins.bottom = 0.75
        worksheet.page_margins.header = 0.3
        worksheet.page_margins.footer = 0.3
    
    def get_suggested_filename(self, project_name: str = None) -> str:
        """
        推奨ファイル名を生成
        
        Args:
            project_name: プロジェクト名
            
        Returns:
            str: 推奨ファイル名
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if project_name:
            # プロジェクト名をファイル名に適した形式に変換
            safe_project_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
            safe_project_name = safe_project_name.replace(' ', '_')
            return f"asana_tasks_{safe_project_name}_{timestamp}.xlsx"
        else:
            return f"asana_tasks_{timestamp}.xlsx"
    
    def validate_output_path(self, filepath: str) -> tuple[bool, str]:
        """
        出力パスの妥当性を検証
        
        Args:
            filepath: 出力ファイルパス
            
        Returns:
            tuple[bool, str]: (有効かどうか, エラーメッセージ)
        """
        try:
            path = Path(filepath)
            
            # 親ディレクトリの存在確認
            if not path.parent.exists():
                try:
                    path.parent.mkdir(parents=True, exist_ok=True)
                except OSError as e:
                    return False, f"ディレクトリの作成に失敗しました: {str(e)}"
            
            # 書き込み権限の確認
            if path.exists():
                if not os.access(path, os.W_OK):
                    return False, "ファイルに書き込み権限がありません"
            else:
                # 親ディレクトリの書き込み権限確認
                if not os.access(path.parent, os.W_OK):
                    return False, "ディレクトリに書き込み権限がありません"
            
            return True, ""
            
        except Exception as e:
            return False, f"パスの検証に失敗しました: {str(e)}"
    
    def _update_progress(self, progress: int) -> None:
        """
        進捗状況を更新
        
        Args:
            progress: 進捗率（0-100）
        """
        if self.progress_callback:
            self.progress_callback(min(max(progress, 0), 100))
    
    def _update_status(self, status: str) -> None:
        """
        ステータスメッセージを更新
        
        Args:
            status: ステータスメッセージ
        """
        self._current_operation = status
        if self.status_callback:
            self.status_callback(status)
        logger.info(f"Excel出力ステータス: {status}")
    
    def get_current_operation(self) -> str:
        """
        現在の操作状況を取得
        
        Returns:
            str: 現在の操作状況
        """
        return self._current_operation
    
    def estimate_processing_time(self, task_count: int, field_count: int) -> int:
        """
        処理時間を推定
        
        Args:
            task_count: タスク数
            field_count: フィールド数
            
        Returns:
            int: 推定処理時間（秒）
        """
        # 基本処理時間（タスク1件あたり約0.01秒）
        base_time = task_count * 0.01
        
        # フィールド数による補正
        field_factor = 1 + (field_count - 5) * 0.1
        
        # フォーマット処理時間
        format_time = task_count * 0.005
        
        # 総処理時間
        total_time = (base_time * field_factor) + format_time
        
        # 最小1秒、最大300秒（5分）
        return max(1, min(int(total_time), 300))
    
    def get_memory_usage_estimate(self, task_count: int, field_count: int) -> int:
        """
        メモリ使用量を推定
        
        Args:
            task_count: タスク数
            field_count: フィールド数
            
        Returns:
            int: 推定メモリ使用量（MB）
        """
        # タスク1件あたりの平均メモリ使用量（KB）
        per_task_memory = field_count * 0.5  # フィールドあたり0.5KB
        
        # 総メモリ使用量（MB）
        total_memory = (task_count * per_task_memory) / 1024
        
        # Excel処理のオーバーヘッド
        overhead = max(10, total_memory * 0.3)
        
        return int(total_memory + overhead)